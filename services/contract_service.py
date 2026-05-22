#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合同业务服务层
"""

from typing import List, Dict, Optional, Tuple
from datetime import datetime
import openpyxl
import csv

from models.database import DatabaseManager
from models.entities import Contract
from utils.helpers import normalize_header, generate_hash, get_days_until_deadline, get_warning_level


class ContractService:
    """合同业务服务"""
    
    def __init__(self):
        self.db = DatabaseManager()
    
    def add_contract(self, contract: Contract) -> Tuple[bool, str]:
        """添加合同"""
        return self.db.add_contract(contract)
    
    def update_contract(self, contract_no: str, data: Dict) -> bool:
        """更新合同"""
        return self.db.update_contract(contract_no, data)
    
    def delete_contract(self, contract_no: str) -> bool:
        """删除合同"""
        return self.db.delete_contract(contract_no)
    
    def get_contract(self, contract_no: str) -> Optional[Contract]:
        """获取单个合同"""
        return self.db.get_contract_by_no(contract_no)
    
    def get_all_contracts(self, filters: Optional[Dict] = None) -> List[Contract]:
        """获取所有合同"""
        return self.db.get_contracts(filters)
    
    def get_warning_contracts(self) -> List[Dict]:
        """获取预警合同"""
        contracts = self.db.get_contracts()
        warning_list = []
        
        for contract in contracts:
            if not contract.合同终止日期:
                continue
            
            days = get_days_until_deadline(contract.合同终止日期)
            if days is None:
                continue
            
            level = get_warning_level(days)
            if level != 'none':
                warning_list.append({
                    'contract': contract,
                    'days': days,
                    'level': level
                })
        
        # 按天数排序
        warning_list.sort(key=lambda x: x['days'])
        return warning_list
    
    def import_from_excel(self, file_path: str) -> Tuple[int, int, List[str]]:
        """从 Excel 导入合同"""
        success_count = 0
        fail_count = 0
        errors = []
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 读取表头并标准化
                headers = []
                for cell in ws[1]:
                    header = str(cell.value).strip() if cell.value else ''
                    normalized = normalize_header(header)
                    headers.append(normalized)
                
                # 读取数据行
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    data = {}
                    for idx, value in enumerate(row):
                        if idx < len(headers):
                            data[headers[idx]] = value
                    
                    # 检查合同编号
                    if not data.get('合同编号'):
                        errors.append(f"第{row_idx}行: 合同编号为空")
                        fail_count += 1
                        continue
                    
                    # 创建合同对象
                    contract = Contract.from_dict(data)
                    
                    # 添加到数据库
                    success, msg = self.add_contract(contract)
                    if success:
                        success_count += 1
                    else:
                        fail_count += 1
                        if "重复" not in msg:
                            errors.append(f"第{row_idx}行: {msg}")
            
            return success_count, fail_count, errors
        
        except Exception as e:
            return 0, 0, [f"导入失败: {str(e)}"]
    
    def export_to_excel(self, file_path: str, contracts: List[Contract]) -> bool:
        """导出合同到 Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "合同数据"
            
            # 写入表头
            from config import CONTRACT_FIELDS
            for col, field in enumerate(CONTRACT_FIELDS, start=1):
                ws.cell(row=1, column=col, value=field)
            
            # 写入数据
            for row_idx, contract in enumerate(contracts, start=2):
                data = contract.to_dict()
                for col, field in enumerate(CONTRACT_FIELDS, start=1):
                    ws.cell(row=row_idx, column=col, value=data.get(field))
            
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"导出失败: {e}")
            return False
    
    def get_statistics(self, contracts: List[Contract]) -> Dict:
        """获取统计数据"""
        if not contracts:
            return {
                'total_count': 0,
                'total_amount': 0,
                'avg_amount': 0,
                'by_region': {},
                'by_salesperson': {},
                'by_year': {}
            }
        
        total_amount = 0
        by_region = {}
        by_salesperson = {}
        by_year = {}
        
        for contract in contracts:
            amount = contract.合同额 or 0
            total_amount += amount
            
            # 按区域统计
            region = contract.区域 or '未知'
            by_region[region] = by_region.get(region, 0) + amount
            
            # 按销售负责人统计
            salesperson = contract.销售负责人 or '未知'
            by_salesperson[salesperson] = by_salesperson.get(salesperson, 0) + amount
            
            # 按年份统计
            if contract.合同签字日期:
                year = contract.合同签字日期[:4] if len(contract.合同签字日期) >= 4 else '未知'
                by_year[year] = by_year.get(year, 0) + amount
        
        return {
            'total_count': len(contracts),
            'total_amount': total_amount,
            'avg_amount': total_amount / len(contracts) if contracts else 0,
            'by_region': by_region,
            'by_salesperson': by_salesperson,
            'by_year': by_year
        }
    
    def get_distinct_regions(self) -> List[str]:
        """获取所有区域"""
        return self.db.get_distinct_values('区域')
    
    def get_distinct_salespersons(self) -> List[str]:
        """获取所有销售负责人"""
        return self.db.get_distinct_values('销售负责人')
