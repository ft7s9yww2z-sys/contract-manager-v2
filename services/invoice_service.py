#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票业务服务层
"""

from typing import List, Dict, Optional, Tuple
import openpyxl

from models.database import DatabaseManager
from models.entities import Invoice
from utils.helpers import normalize_header, generate_hash


class InvoiceService:
    """发票业务服务"""
    
    def __init__(self):
        self.db = DatabaseManager()
    
    def add_invoice(self, invoice: Invoice) -> Tuple[bool, str]:
        """添加发票"""
        return self.db.add_invoice(invoice)
    
    def get_all_invoices(self, filters: Optional[Dict] = None) -> List[Invoice]:
        """获取所有发票"""
        return self.db.get_invoices(filters)
    
    def import_from_excel(self, file_path: str) -> Tuple[int, int, List[str]]:
        """从 Excel 导入发票"""
        success_count = 0
        fail_count = 0
        errors = []
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 读取表头并标准化
                headers = []
                for cell in ws[1]:
                    header = str(cell.value).strip() if cell.value else ''
                    headers.append(header)
                
                # 读取数据行
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    data = {}
                    for idx, value in enumerate(row):
                        if idx < len(headers):
                            data[headers[idx]] = value
                    
                    # 检查合同号
                    if not data.get('合同号'):
                        errors.append(f"第{row_idx}行: 合同号为空")
                        fail_count += 1
                        continue
                    
                    # 创建发票对象
                    invoice = Invoice(
                        开票日期=data.get('开票日期'),
                        合同号=data.get('合同号', ''),
                        付款单位名称=data.get('付款单位名称'),
                        代码=data.get('代码'),
                        发票金额=data.get('发票金额'),
                        发票项目=data.get('发票项目'),
                        类型=data.get('类型'),
                        发票类型=data.get('发票类型'),
                        除税=data.get('除税'),
                        备注=data.get('备注')
                    )
                    
                    # 添加到数据库
                    success, msg = self.add_invoice(invoice)
                    if success:
                        success_count += 1
                    else:
                        fail_count += 1
                        if "重复" not in msg:
                            errors.append(f"第{row_idx}行: {msg}")
            
            return success_count, fail_count, errors
        
        except Exception as e:
            return 0, 0, [f"导入失败: {str(e)}"]
    
    def export_to_excel(self, file_path: str, invoices: List[Invoice]) -> bool:
        """导出发票到 Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "发票数据"
            
            # 写入表头
            from config import INVOICE_FIELDS
            for col, field in enumerate(INVOICE_FIELDS, start=1):
                ws.cell(row=1, column=col, value=field)
            
            # 写入数据
            for row_idx, invoice in enumerate(invoices, start=2):
                data = invoice.to_dict()
                for col, field in enumerate(INVOICE_FIELDS, start=1):
                    ws.cell(row=row_idx, column=col, value=data.get(field))
            
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"导出失败: {e}")
            return False
    
    def get_statistics(self, invoices: List[Invoice]) -> Dict:
        """获取发票统计数据"""
        if not invoices:
            return {
                'total_count': 0,
                'total_amount': 0,
                'avg_amount': 0
            }
        
        total_amount = sum(inv.发票金额 or 0 for inv in invoices)
        
        return {
            'total_count': len(invoices),
            'total_amount': total_amount,
            'avg_amount': total_amount / len(invoices) if invoices else 0
        }
